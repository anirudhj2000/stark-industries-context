#!/usr/bin/env python3
"""
Excel/Spreadsheet Extraction Helper Script

Extracts data from Excel files (XLS, XLSX) including formulas and values.
"""
import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime, date, time

try:
    import openpyxl
    from openpyxl.cell.cell import TYPE_FORMULA, TYPE_NUMERIC, TYPE_STRING, TYPE_BOOL
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def extract_cell_info(cell) -> Dict[str, Any]:
    """Extract complete information from a cell including formula and value"""
    # Convert datetime objects to ISO format strings for JSON serialization
    value = cell.value
    if isinstance(value, (datetime, date)):
        value = value.isoformat()
    elif isinstance(value, time):
        value = value.isoformat()

    cell_info = {
        "value": value,
        "formula": None,
        "data_type": cell.data_type,
        "is_formula": False
    }

    # Check if cell contains a formula
    if cell.data_type == TYPE_FORMULA:
        cell_info["formula"] = cell.value  # The formula string
        cell_info["is_formula"] = True
        # Try to get calculated value (may be None if workbook wasn't calculated)
        try:
            calc_val = cell.value if hasattr(cell, '_value') else None
            if isinstance(calc_val, (datetime, date, time)):
                calc_val = calc_val.isoformat()
            cell_info["calculated_value"] = calc_val
        except:
            cell_info["calculated_value"] = None

    # Format value as string for JSON serialization
    if cell_info["value"] is not None:
        cell_info["display_value"] = str(cell_info["value"])
    else:
        cell_info["display_value"] = ""

    return cell_info


def extract_excel(file_path: Path, include_formulas: bool = True) -> Dict[str, Any]:
    """Extract content from Excel file with formulas and values"""
    try:
        # Load with data_only=False to preserve formulas
        workbook = openpyxl.load_workbook(file_path, data_only=False)

        sheets = []
        total_rows = 0
        total_cells = 0
        total_formulas = 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column

            # Extract detailed cell data
            data = []
            formula_count = 0

            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row,
                                                          min_col=1, max_col=max_col)):
                row_data = []
                for cell in row:
                    if include_formulas:
                        cell_info = extract_cell_info(cell)
                        if cell_info["is_formula"]:
                            formula_count += 1
                        row_data.append(cell_info)
                    else:
                        # Simple mode: just values
                        row_data.append(str(cell.value) if cell.value is not None else "")

                data.append(row_data)

            # Detect if first row is header
            has_header = False
            if data and len(data) > 1:
                first_row = data[0]
                if include_formulas:
                    # Check if first row cells are non-empty strings
                    has_header = any(
                        cell.get("display_value", "").strip()
                        for cell in first_row
                        if isinstance(cell, dict)
                    )
                else:
                    has_header = any(isinstance(cell, str) and cell.strip() for cell in first_row)

            # Extract headers
            headers = []
            if has_header and data:
                if include_formulas:
                    headers = [cell.get("display_value", "") for cell in data[0]]
                else:
                    headers = data[0]

            # Build records format
            records = []
            if has_header and len(data) > 1:
                for row_data in data[1:]:
                    record = {}
                    for i, cell_data in enumerate(row_data):
                        header = headers[i] if i < len(headers) else f"Column_{i+1}"
                        record[header] = cell_data
                    records.append(record)

            sheet_info = {
                "name": sheet_name,
                "rows": max_row,
                "columns": max_col,
                "has_header": has_header,
                "headers": headers,
                "data": data,
                "records": records if has_header else [],
                "formula_count": formula_count
            }

            sheets.append(sheet_info)
            total_rows += max_row
            total_cells += max_row * max_col
            total_formulas += formula_count

        # Build result
        result = {
            "file_name": file_path.name,
            "file_size_bytes": file_path.stat().st_size,
            "format": file_path.suffix.upper().replace(".", ""),
            "sheet_count": len(sheets),
            "sheets": sheets,
            "total_rows": total_rows,
            "total_cells": total_cells,
            "total_formulas": total_formulas,
            "metadata": {
                "engine": "openpyxl",
                "includes_formulas": include_formulas,
                "workbook_properties": {
                    "creator": workbook.properties.creator if workbook.properties else None,
                    "title": workbook.properties.title if workbook.properties else None,
                    "subject": workbook.properties.subject if workbook.properties else None,
                    "created": str(workbook.properties.created) if workbook.properties and workbook.properties.created else None,
                    "modified": str(workbook.properties.modified) if workbook.properties and workbook.properties.modified else None
                }
            }
        }

        return result

    except Exception as e:
        raise Exception(f"Failed to extract Excel file: {e}")


def main():
    parser = argparse.ArgumentParser(description='Extract data from Excel files')
    parser.add_argument('--input', required=True, help='Path to Excel file')
    parser.add_argument('--output', required=True, help='Output JSON file')
    parser.add_argument('--include-formulas', action='store_true', default=True,
                       help='Include formulas (default: True)')
    parser.add_argument('--no-formulas', action='store_true',
                       help='Exclude formulas, values only')

    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Excel file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    try:
        include_formulas = not args.no_formulas
        result = extract_excel(input_path, include_formulas=include_formulas)

        # Write output
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)

        print(f"SUCCESS: Extracted {result['sheet_count']} sheets from {input_path.name}")
        print(f"Total rows: {result['total_rows']}, Total cells: {result['total_cells']}")
        if include_formulas:
            print(f"Total formulas: {result['total_formulas']}")

    except Exception as e:
        print(f"ERROR: Extraction failed: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
